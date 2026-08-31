from collections import Counter, defaultdict
from datetime import datetime, date
from itertools import combinations
from pathlib import Path
import json
import re

from openpyxl import load_workbook

try:
    import jieba  # type: ignore
except Exception:
    jieba = None


WORKBOOK = next(Path(".").glob("*.xlsx"))
OUT = Path("data.js")


STOPWORDS = {
    "这个", "那个", "他们", "我们", "你们", "她们", "他们", "自己", "就是", "还是",
    "一个", "没有", "不是", "因为", "所以", "可以", "已经", "真的", "非常", "比较",
    "感觉", "效果", "商品", "产品", "包装", "物流", "发货", "收到", "使用", "购买",
    "评论", "评价", "好评", "差评", "中评", "不错", "很好", "一般", "还行", "挺好",
    "同仁堂", "金匮肾气丸", "肾气丸", "药品", "药丸", "正品", "店家", "客服", "快递",
}

SCENE_RULES = [
    ("疼痛缓解", ["疼", "痛", "止痛", "缓解", "腰", "膝", "关节", "肩", "颈"]),
    ("睡眠改善", ["睡眠", "失眠", "入睡", "夜尿", "睡得", "精神"]),
    ("物流发货", ["物流", "发货", "快递", "到货", "收货", "送货"]),
    ("包装体验", ["包装", "外观", "盒子", "精致", "拆包", "密封"]),
    ("价格优惠", ["价格", "便宜", "优惠", "活动", "划算", "性价比"]),
    ("品质正品", ["正品", "品质", "品牌", "药效", "真品", "药房"]),
    ("客服服务", ["客服", "服务", "态度", "耐心", "咨询"]),
    ("复购回购", ["复购", "回购", "再买", "继续", "还会", "多买"]),
]

BAD_RULES = [
    ("效果不明", ["没效果", "无效", "不明显", "没感觉", "效果一般", "不管用"]),
    ("假货质疑", ["假", "上当", "智商税", "不是正品", "踩雷"]),
    ("物流问题", ["物流慢", "快递慢", "延迟", "漏发", "发错", "少发"]),
    ("包装问题", ["包装差", "破损", "脏", "粗糙"]),
    ("客服问题", ["客服差", "态度差", "不处理", "不回复"]),
    ("口感不佳", ["太苦", "难吃", "异味", "怪味"]),
    ("价格偏高", ["贵", "不值", "太贵"]),
]

FOCUS_RULES = [
    ("效果", ["效果", "缓解", "改善", "止痛", "见效"]),
    ("疼痛", ["疼", "痛", "腰", "膝", "关节", "肩", "颈"]),
    ("睡眠", ["睡眠", "失眠", "入睡", "夜尿"]),
    ("物流", ["物流", "发货", "快递", "到货"]),
    ("包装", ["包装", "外观", "盒子", "精致"]),
    ("价格", ["价格", "便宜", "优惠", "活动"]),
    ("正品", ["正品", "品牌", "药房"]),
    ("复购", ["复购", "回购", "再买", "继续"]),
]

DOMAIN_TERMS = [
    "效果", "缓解", "止痛", "疼痛", "腰痛", "膝痛", "关节", "睡眠", "失眠", "入睡", "夜尿",
    "物流", "发货", "快递", "包装", "外观", "正品", "品牌", "客服", "服务", "价格", "优惠",
    "便宜", "活动", "回购", "复购", "疗程", "体验", "推荐", "口感", "药效", "中药", "药品",
    "满意", "方便", "携带", "到货", "收货", "疼痛缓解", "质量", "药房", "正品保障", "药效明显",
]


def text(v):
    return "" if v is None else str(v).strip()


def norm_month(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m")
    s = text(v)
    return s[:7] if len(s) >= 7 else s


def tokenize(s):
    s = text(s)
    if not s:
        return []
    if jieba is not None:
        words = jieba.lcut(s)
    else:
        words = re.findall(r"[\u4e00-\u9fff]{2,}|[A-Za-z0-9.+*/]+", s)
    out = []
    for w in words:
        w = w.strip()
        if len(w) < 2:
            continue
        if w in STOPWORDS:
            continue
        if re.fullmatch(r"[0-9.]+", w):
            continue
        out.append(w)
    return out


def matches_any(s, rules):
    return [name for name, kws in rules if any(k in s for k in kws)]


def sku_group(sku):
    if "疗程装" in sku:
        return "疗程装"
    if "体验装" in sku:
        return "体验装"
    if "推荐装" in sku:
        return "推荐装"
    if "促销" in sku or "优惠" in sku or "大促" in sku:
        return "促销装"
    if "CCTV" in sku:
        return "CCTV推荐"
    if "1盒" in sku:
        return "1盒装"
    if "2盒" in sku:
        return "2盒装"
    return "其他"


wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]

rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    rows.append(
        {
            "idx": r[0],
            "user": text(r[1]),
            "date": r[2],
            "type": text(r[3]),
            "sku": text(r[4]),
            "review": text(r[5]),
            "helpful": r[6] or 0,
            "reply": text(r[7]),
            "reply_date": r[8],
        }
    )

total = len(rows)
valid = sum(1 for r in rows if r["review"])
counts = Counter(r["type"] for r in rows)
good = counts.get("好评", 0)
mid = counts.get("中评", 0)
bad = counts.get("差评", 0)
real_total = good + mid + bad
template = 0

monthly = defaultdict(lambda: Counter())
weekday_stats = defaultdict(Counter)
sku_stats = defaultdict(lambda: Counter())
sku_type_stats = defaultdict(lambda: Counter())
sku_name_counts = Counter()
reply_monthly = Counter()

good_comments = []
bad_comments = []
repurchase_samples = []
scene_counts = Counter()
focus_counts = Counter()
cooccur = Counter()
reply_count = 0

keyword_good = Counter()
keyword_bad = Counter()

for r in rows:
    month = norm_month(r["date"])
    monthly[month][r["type"]] += 1
    if isinstance(r["date"], (datetime, date)):
        weekday_name = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"][r["date"].weekday()]
    else:
        try:
            weekday_name = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"][
                datetime.strptime(text(r["date"])[:10], "%Y-%m-%d").weekday()
            ]
        except ValueError:
            weekday_name = None
    if weekday_name:
        weekday_stats[weekday_name][r["type"]] += 1
    review = r["review"]
    sku = r["sku"]
    sku_name_counts[sku] += 1
    sku_stats[sku][r["type"]] += 1
    sku_type_stats[sku_group(sku)][r["type"]] += 1
    if r["reply"]:
        reply_count += 1
        rep_month = norm_month(r["reply_date"] or r["date"])
        reply_monthly[rep_month] += 1
        if any(k in review for k in ["复购", "回购", "再买", "继续", "还会"]):
            repurchase_samples.append(review)
    if r["type"] == "好评":
        good_comments.append(review)
    elif r["type"] == "差评":
        bad_comments.append(review)

    for name, kws in SCENE_RULES:
        if any(k in review for k in kws):
            scene_counts[name] += 1
    for name, kws in FOCUS_RULES:
        if any(k in review for k in kws):
            focus_counts[name] += 1
    matched_focus = [name for name, kws in FOCUS_RULES if any(k in review for k in kws)]
    for a, b in combinations(sorted(set(matched_focus)), 2):
        cooccur[(a, b)] += 1

    for tok in tokenize(review):
        if r["type"] == "差评":
            keyword_bad[tok] += 1
        else:
            keyword_good[tok] += 1

def top_samples(rules, source_rows, limit=8):
    out = []
    for name, kws in rules:
        samples = []
        count = 0
        for row in source_rows:
            if any(k in row["review"] for k in kws):
                count += 1
                if len(samples) < 3 and row["review"] not in samples:
                    samples.append(row["review"])
        if count:
            out.append({"theme": name, "count": count, "rate": round(count / max(real_total, 1) * 100, 1), "samples": samples})
    out.sort(key=lambda x: x["count"], reverse=True)
    return out[:limit]


def top_bad(rules, source_rows, limit=7):
    out = []
    for name, kws in rules:
        samples = []
        count = 0
        for row in source_rows:
            if row["type"] in {"差评", "中评"} and any(k in row["review"] for k in kws):
                count += 1
                if len(samples) < 3 and row["review"] not in samples:
                    samples.append(row["review"])
        if count:
            out.append({"reason": name, "count": count, "rate": round(count / max(real_total, 1) * 100, 1), "samples": samples})
    out.sort(key=lambda x: x["count"], reverse=True)
    return out[:limit]


monthly_list = []
for month in sorted(monthly):
    c = monthly[month]
    total_m = c["好评"] + c["中评"] + c["差评"]
    monthly_list.append(
        {
            "month": month + "-01",
            "count": total_m,
            "template": 0,
            "good": c["好评"],
            "mid": c["中评"],
            "bad": c["差评"],
            "real_rate": round((c["好评"] / total_m * 100) if total_m else 0, 1),
        }
    )

weekday_list = [
    {
        "name": name,
        "value": sum(weekday_stats[name].values()),
        "good": weekday_stats[name]["好评"],
        "mid": weekday_stats[name]["中评"],
        "bad": weekday_stats[name]["差评"],
    }
    for name in ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
]

sku_dist = [{"name": sku, "value": cnt} for sku, cnt in sku_name_counts.most_common(10)]

sku_rate = []
for sku, cnt in sku_name_counts.most_common():
    st = sku_stats[sku]
    sku_rate.append(
        {
            "sku": sku,
            "total": cnt,
            "real_rate": round((st["好评"] / cnt * 100) if cnt else 0, 1),
            "bad_n": st["差评"],
        }
    )

top_skus = [sku for sku, _ in sku_name_counts.most_common(8)]
sku_heat = {
    "skus": top_skus,
    "types": ["好评", "中评", "差评"],
    "matrix": [[sku_type_stats[sku_group(sku)][t] for t in ["好评", "中评", "差评"]] for sku in top_skus],
}

spec_type = Counter(sku_group(r["sku"]) for r in rows)
scene_list = [{"name": name, "count": cnt} for name, cnt in scene_counts.most_common(6)]
focus_list = [{"name": name, "count": cnt} for name, cnt in focus_counts.most_common(4)]

good_wordcloud = [{"name": k, "value": v} for k, v in (keyword_good + Counter({t: 0 for t in DOMAIN_TERMS})).most_common(40) if v > 0]
bad_wordcloud = [{"name": k, "value": v} for k, v in (keyword_bad + Counter({t: 0 for t in DOMAIN_TERMS})).most_common(35) if v > 0]

promo_terms = Counter()
for sku, cnt in sku_name_counts.items():
    for part in re.findall(r"[\u4e00-\u9fff]{2,}|[0-9.]+(?:mg|ml|g|贴|盒|丸)?", sku):
        if len(part) < 2:
            continue
        if part in STOPWORDS:
            continue
        promo_terms[part] += cnt
promo = [{"name": k, "value": v} for k, v in promo_terms.most_common(8)]

repurchase_trend = [{"month": m + "-01", "count": reply_monthly[m]} for m in sorted(reply_monthly)]
repurchase_samples = repurchase_samples[:6] or [r["review"] for r in rows if r["reply"]][:6]

cooccur_list = [
    {"good": a, "bad": b, "count": c}
    for (a, b), c in cooccur.most_common(6)
]

data = {
    "meta": {
        "source": WORKBOOK.name,
        "generated": datetime.now().strftime("%Y-%m-%d"),
        "total": total,
        "valid": valid,
    },
    "kpi": {
        "total": total,
        "template": template,
        "template_rate": round(template / total * 100, 1) if total else 0,
        "good": good,
        "mid": mid,
        "bad": bad,
        "real_total": real_total,
        "real_good_rate": round(good / real_total * 100, 1) if real_total else 0,
        "overall_good_rate": round(good / total * 100, 1) if total else 0,
        "real_rate": 100.0,
        "repurchase": reply_count,
        "avg_len": round(sum(len(r["review"]) for r in rows) / total, 1) if total else 0,
    },
    "sentiment_pie": [
        {"name": "真实好评", "value": good},
        {"name": "中评", "value": mid},
        {"name": "差评", "value": bad},
        {"name": "模板好评(无内容)", "value": template},
    ],
    "good_themes": top_samples(SCENE_RULES, rows),
    "bad_reasons": top_bad(BAD_RULES, rows),
    "monthly": monthly_list,
    "weekday": weekday_list,
    "wordcloud": {"good": good_wordcloud, "bad": bad_wordcloud},
    "sku_dist": sku_dist,
    "sku_heat": sku_heat,
    "promo": promo,
    "repurchase": {
        "trend": repurchase_trend,
        "samples": repurchase_samples,
    },
    "cooccur": cooccur_list,
    "sku_rate": sku_rate,
    "portrait": {
        "spec_type": [{"name": k, "value": v} for k, v in spec_type.most_common()],
        "scene": scene_list,
        "focus": focus_list,
    },
}

OUT.write_text("window.DATA = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n", encoding="utf-8")
print(f"Wrote {OUT} from {WORKBOOK}")
